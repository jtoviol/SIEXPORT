"""Repository de Vacunación — lee el Excel uploadeado, NO consulta SQL.

A diferencia de los otros 4 módulos, Vacunación no toca AVS_REGISTROS_SERAGIL.
El operador sube un .xlsx con la estructura que ya vimos (32 columnas, mismo
shape que la query DI pero solo programas de vacunación) y este repository lo
parsea fila a fila.

El régimen viene en la columna `REGIMEN` del Excel — no hay cruce contra
AVS_REGISTROS_AP ni se usa el filtro CAB/FAB.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Protocol

from efdi.domain.models import (
    RegistroVacuna,
    Regimen,
    Sexo,
    TipoDocumento,
)

log = logging.getLogger(__name__)


# Columnas mínimas que el Excel DEBE tener. Si falta alguna, error explícito al
# subir. Las demás son opcionales y se rellenan con None si no vienen.
COLUMNAS_REQUERIDAS: set[str] = {
    "SEQ_SERAGIL",
    "REGIMEN",
    "AFL_PRIMER_NOMBRE",
    "AFL_PRIMER_APELLIDO",
    "DES_TIPO_IDENTIFICACION",
    "NRO_TIPO_IDENTIFICACION",
    "DES_PROGRAMA_DEMIND",
    "FEC_REGISTRO_INFORMACION",
    "FEC_NACIMIENTO_PERSONA",
    "VLR_EDAD_ACTUAL",
    "DES_GENERO",
}


class VacunacionRepository(Protocol):
    """Contrato del repository."""

    def get_total(self, excel_path: Path, regimen: str | None = None) -> int: ...

    def obtener_registros(
        self,
        excel_path: Path,
        regimen: str | None = None,
        limite: int = 0,
        offset: int = 0,
    ) -> list[RegistroVacuna]: ...

    def resumen(self, excel_path: Path) -> dict: ...


# ─── Helpers de normalización ───────────────────────────────────────────────


def _normalizar_sexo(des: str | None) -> Sexo:
    """DES_GENERO viene como 'MASCULINO'/'FEMENINO'."""
    if des:
        u = des.upper().strip()
        if u.startswith("F") or "FEM" in u:
            return Sexo.F
        if u.startswith("M") or "MASC" in u:
            return Sexo.M
    return Sexo.M


def _normalizar_tipo_doc(des: str | None) -> TipoDocumento:
    """DES_TIPO_IDENTIFICACION viene como 'CEDULA DE CIUDADANIA', 'TARJETA DE IDENTIDAD', etc."""
    if not des:
        return TipoDocumento.CC
    u = des.upper().strip()
    if "CIUDAD" in u or u == "CC":
        return TipoDocumento.CC
    if "TARJETA" in u or u == "TI":
        return TipoDocumento.TI
    if "REGISTRO" in u or u == "RC":
        return TipoDocumento.RC
    if "EXTRANJ" in u or u == "CE":
        return TipoDocumento.CE
    if "PASAPORTE" in u or u == "PA":
        return TipoDocumento.PA
    if "MENOR SIN ID" in u or "MSI" in u or u == "MS":
        return TipoDocumento.MS
    return TipoDocumento.CC


def _normalizar_regimen(s: str | None) -> Regimen | None:
    if not s:
        return None
    u = s.upper().strip()
    if u == "CONTRIBUTIVO":
        return Regimen.CONTRIBUTIVO
    if u == "SUBSIDIADO":
        return Regimen.SUBSIDIADO
    if u == "VINCULADO":
        return Regimen.VINCULADO
    return None


def _parse_date(v: object) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _str_or_none(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.upper() == "NULL" or s.upper() == "NOTIENE":
        return None
    return s


def _int_or_none(v: object) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


# ─── Loader del Excel ───────────────────────────────────────────────────────


def _open_workbook(excel_path: Path):
    """Abre el .xlsx en modo read_only (no carga todo en memoria)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl no instalado. Instalar con: pip install '.[excel]'"
        ) from e
    return load_workbook(excel_path, read_only=True, data_only=True)


def _validar_columnas(headers: list[str], excel_path: Path) -> dict[str, int]:
    """Verifica que todas las columnas requeridas estén presentes.
    Devuelve mapa {nombre_columna: indice}.
    """
    idx = {h: i for i, h in enumerate(headers) if h}
    faltantes = COLUMNAS_REQUERIDAS - set(idx.keys())
    if faltantes:
        raise ValueError(
            f"El Excel '{excel_path.name}' no tiene las columnas requeridas: "
            f"{sorted(faltantes)}. Columnas presentes: {sorted(idx.keys())}"
        )
    return idx


def _row_a_registro(row: tuple, idx: dict[str, int]) -> RegistroVacuna | None:
    """Convierte una fila del Excel a RegistroVacuna. Devuelve None si la fila
    es inválida (faltan campos críticos)."""

    def g(col: str) -> object:
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    fec_aplicacion = _parse_date(g("FEC_REGISTRO_INFORMACION"))
    fec_nac = _parse_date(g("FEC_NACIMIENTO_PERSONA"))
    seq = _int_or_none(g("SEQ_SERAGIL"))
    num_doc = _str_or_none(g("NRO_TIPO_IDENTIFICACION"))
    primer_nombre = _str_or_none(g("AFL_PRIMER_NOMBRE"))
    primer_apellido = _str_or_none(g("AFL_PRIMER_APELLIDO"))
    programa = _str_or_none(g("DES_PROGRAMA_DEMIND"))

    if not all([fec_aplicacion, fec_nac, seq, num_doc, primer_nombre, primer_apellido, programa]):
        return None

    try:
        return RegistroVacuna(
            seq_seragil=int(seq),  # type: ignore[arg-type]
            tipo_documento=_normalizar_tipo_doc(_str_or_none(g("DES_TIPO_IDENTIFICACION"))),
            num_documento=str(num_doc),
            tipo_identificacion_desc=_str_or_none(g("DES_TIPO_IDENTIFICACION")),
            primer_nombre=str(primer_nombre),
            segundo_nombre=_str_or_none(g("AFL_SEGUNDO_NOMBRE")),
            primer_apellido=str(primer_apellido),
            segundo_apellido=_str_or_none(g("AFL_SEGUNDO_APELLIDO")),
            sexo=_normalizar_sexo(_str_or_none(g("DES_GENERO"))),
            edad=_int_or_none(g("VLR_EDAD_ACTUAL")) or 0,
            fecha_nacimiento=fec_nac,  # type: ignore[arg-type]
            direccion=_str_or_none(g("DES_DIRECCION_ACTUAL")),
            telefono_1=_str_or_none(g("DES_TELEFONO_UNO")),
            telefono_2=_str_or_none(g("DES_TELEFONO_DOS")),
            correo=_str_or_none(g("DES_CORREO_ELECTRONICO")),
            departamento=_str_or_none(g("DES_DEPARTAMENTO")),
            municipio=_str_or_none(g("DES_MUNICIPIO")),
            zona_afiliado=_int_or_none(g("ZONA_AFILIADO")),
            curso_vida=_str_or_none(g("DES_CURSO_VIDA_ASOCIADO")),
            regimen=_normalizar_regimen(_str_or_none(g("REGIMEN"))),
            fecha_aplicacion=fec_aplicacion,  # type: ignore[arg-type]
            programa=str(programa),
            modo_ingreso=_str_or_none(g("DES_MODO_INGRESO")),
            encuestador=_str_or_none(g("ENCUESTADOR")),
            cargo_encuestador=_str_or_none(g("DES_CARGO_USUARIO")),
        )
    except Exception:
        log.exception("vacunacion.row_invalida seq=%s", seq)
        return None


def _filtro_regimen_matchea(reg: Regimen | None, filtro: str | None) -> bool:
    """`filtro` puede ser 'SUBSIDIADO', 'CONTRIBUTIVO' o None (sin filtro)."""
    if not filtro:
        return True
    f = filtro.upper().strip()
    if reg is None:
        return False
    return str(reg).upper() == f or reg.value.upper() == f


# ─── Implementación ─────────────────────────────────────────────────────────


class ExcelVacunacionRepository:
    """Lee del Excel uploadeado. NO mock — el flujo depende del archivo real."""

    def resumen(self, excel_path: Path) -> dict:
        """Cuenta filas y desglosa por régimen. Útil para el preview post-upload."""
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel no encontrado: {excel_path}")
        wb = _open_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        headers_row = next(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in headers_row]
        idx = _validar_columnas(headers, excel_path)

        total = 0
        por_regimen: dict[str, int] = {"SUBSIDIADO": 0, "CONTRIBUTIVO": 0, "OTRO": 0}
        afiliados_por_regimen: dict[str, set[str]] = {"SUBSIDIADO": set(), "CONTRIBUTIVO": set(), "OTRO": set()}

        col_reg = idx["REGIMEN"]
        col_tipo_doc = idx["DES_TIPO_IDENTIFICACION"]
        col_num_doc = idx["NRO_TIPO_IDENTIFICACION"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(c is None for c in row):
                continue
            total += 1
            reg_val = str(row[col_reg]).upper().strip() if col_reg < len(row) and row[col_reg] else ""
            bucket = reg_val if reg_val in ("SUBSIDIADO", "CONTRIBUTIVO") else "OTRO"
            por_regimen[bucket] += 1
            tipo = row[col_tipo_doc] if col_tipo_doc < len(row) else None
            num = row[col_num_doc] if col_num_doc < len(row) else None
            if tipo and num:
                afiliados_por_regimen[bucket].add(f"{tipo}|{num}")

        wb.close()
        return {
            "total_filas": total,
            "por_regimen": {k: por_regimen[k] for k in ("SUBSIDIADO", "CONTRIBUTIVO", "OTRO")},
            "afiliados_por_regimen": {
                k: len(afiliados_por_regimen[k])
                for k in ("SUBSIDIADO", "CONTRIBUTIVO", "OTRO")
            },
        }

    def get_total(self, excel_path: Path, regimen: str | None = None) -> int:
        """Cuenta filas-programa del Excel filtradas por régimen.
        Es el `limite` que va a usar la paginación: misma unidad que devuelve
        `obtener_registros`.
        """
        if not excel_path.exists():
            return 0
        wb = _open_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        headers_row = next(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in headers_row]
        idx = _validar_columnas(headers, excel_path)

        col_reg = idx["REGIMEN"]
        filtro_upper = regimen.upper().strip() if regimen else None
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(c is None for c in row):
                continue
            if filtro_upper:
                reg_val = str(row[col_reg]).upper().strip() if col_reg < len(row) and row[col_reg] else ""
                if reg_val != filtro_upper:
                    continue
            total += 1
        wb.close()
        return total

    def obtener_registros(
        self,
        excel_path: Path,
        regimen: str | None = None,
        limite: int = 0,
        offset: int = 0,
    ) -> list[RegistroVacuna]:
        """Devuelve hasta `limite` filas a partir de `offset`, ya filtradas por régimen.
        Filas con datos incompletos se descartan silenciosamente (log warning).
        """
        if not excel_path.exists():
            return []
        wb = _open_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        headers_row = next(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in headers_row]
        idx = _validar_columnas(headers, excel_path)

        col_reg = idx["REGIMEN"]
        filtro_upper = regimen.upper().strip() if regimen else None

        registros: list[RegistroVacuna] = []
        salteadas = 0
        descartadas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(c is None for c in row):
                continue
            if filtro_upper:
                reg_val = str(row[col_reg]).upper().strip() if col_reg < len(row) and row[col_reg] else ""
                if reg_val != filtro_upper:
                    continue
            if salteadas < offset:
                salteadas += 1
                continue
            if limite > 0 and len(registros) >= limite:
                break
            r = _row_a_registro(row, idx)
            if r is None:
                descartadas += 1
                continue
            registros.append(r)

        wb.close()
        log.info(
            "vacunacion.fetched",
            extra={
                "rows": len(registros),
                "descartadas": descartadas,
                "regimen": regimen or "all",
                "excel": str(excel_path),
            },
        )
        return registros


def get_vacunacion_repository() -> VacunacionRepository:
    """Por ahora solo hay implementación de Excel. No hay mock — para probar
    el flujo, dejar un .xlsx de muestra en data/uploads/vacunacion/."""
    return ExcelVacunacionRepository()
